from __future__ import annotations

import csv
import hashlib
import io
import re
import shutil
import tempfile
import unicodedata
import zipfile
from contextlib import redirect_stdout
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Generator, Iterable, List, Optional, Set, Tuple

import streamlit as st

# ==========================================================
# Ceifador 1 — Lotes por Matriz (XML + ZIP dentro de ZIP)
# ==========================================================


def c1_chave_natural(texto: str) -> List[object]:
    partes = re.split(r"(\d+)", texto.lower())
    return [int(p) if p.isdigit() else p for p in partes]


def c1_limpar_nome_arquivo(texto: str) -> str:
    nome = re.sub(r'[\\/:*?"<>|]+', "_", texto).strip()
    nome = re.sub(r"\s+", " ", nome)
    return nome or "matriz"


def c1_obter_matriz_do_relativo(relativo: Path, fallback: str) -> str:
    partes = relativo.parts
    if not partes:
        return fallback
    primeiro = partes[0]
    return Path(primeiro).stem if primeiro.lower().endswith(".zip") else primeiro


def c1_iterar_xmls_em_zip(
    zip_bytes: bytes, caminho_contexto: str, matriz_nome: str
) -> Generator[Tuple[str, str, str, bytes], None, None]:
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue

            nome_interno = info.filename
            nome_minusculo = nome_interno.lower()

            try:
                conteudo = zf.read(info)
            except Exception as exc:
                print(f"[AVISO] Falha ao ler '{caminho_contexto}!{nome_interno}': {exc}")
                continue

            if nome_minusculo.endswith(".xml"):
                nome_sugerido = Path(nome_interno).name
                chave_ordenacao = f"{caminho_contexto}!{nome_interno}"
                yield matriz_nome, chave_ordenacao, nome_sugerido, conteudo
            elif nome_minusculo.endswith(".zip"):
                sub_contexto = f"{caminho_contexto}!{nome_interno}"
                yield from c1_iterar_xmls_em_zip(conteudo, sub_contexto, matriz_nome)


def c1_iterar_xmls_da_pasta(pasta_raiz: Path) -> Generator[Tuple[str, str, str, bytes], None, None]:
    for caminho in pasta_raiz.rglob("*"):
        if not caminho.is_file():
            continue

        nome_minusculo = caminho.name.lower()

        if nome_minusculo.endswith(".xml"):
            try:
                conteudo = caminho.read_bytes()
            except Exception as exc:
                print(f"[AVISO] Falha ao ler XML '{caminho}': {exc}")
                continue

            relativo = caminho.relative_to(pasta_raiz)
            matriz_nome = c1_obter_matriz_do_relativo(relativo, pasta_raiz.name)
            yield matriz_nome, str(relativo), caminho.name, conteudo

        elif nome_minusculo.endswith(".zip"):
            try:
                zip_bytes = caminho.read_bytes()
            except Exception as exc:
                print(f"[AVISO] Falha ao ler ZIP '{caminho}': {exc}")
                continue

            relativo = caminho.relative_to(pasta_raiz)
            matriz_nome = c1_obter_matriz_do_relativo(relativo, caminho.stem)
            yield from c1_iterar_xmls_em_zip(zip_bytes, str(caminho), matriz_nome)


def c1_gerar_nome_unico(nome_base: str, existentes: Set[str]) -> str:
    if nome_base not in existentes:
        existentes.add(nome_base)
        return nome_base
    stem = Path(nome_base).stem
    suffix = Path(nome_base).suffix
    i = 1
    while True:
        candidato = f"{stem}_{i}{suffix}"
        if candidato not in existentes:
            existentes.add(candidato)
            return candidato
        i += 1


def c1_chunks(iteravel: Iterable[Tuple[str, bytes]], tamanho: int):
    lote: List[Tuple[str, bytes]] = []
    for item in iteravel:
        lote.append(item)
        if len(lote) >= tamanho:
            yield lote
            lote = []
    if lote:
        yield lote


def ceifador1_criar_lotes_zip(pasta_raiz: Path, max_xmls_por_lote: int = 5000, prefixo_lote: str = "lote") -> None:
    print(f"[INFO] Lendo XMLs em '{pasta_raiz}'...")
    por_matriz: Dict[str, List[Tuple[str, str, bytes]]] = {}

    for matriz_nome, chave_ord, nome_xml, conteudo in c1_iterar_xmls_da_pasta(pasta_raiz):
        por_matriz.setdefault(matriz_nome, []).append((chave_ord, nome_xml, conteudo))

    total_xmls = 0
    total_lotes = 0

    for matriz_nome in sorted(por_matriz.keys(), key=c1_chave_natural):
        itens = por_matriz[matriz_nome]
        itens.sort(key=lambda x: c1_chave_natural(x[0]))
        nomes_matriz: Set[str] = set()
        matriz_limpa = c1_limpar_nome_arquivo(matriz_nome)

        for idx_lote, lote in enumerate(
            c1_chunks(((nome_xml, conteudo) for _, nome_xml, conteudo in itens), max_xmls_por_lote),
            start=1,
        ):
            caminho_lote = pasta_raiz / f"{matriz_limpa}_{prefixo_lote}{idx_lote}.zip"
            with zipfile.ZipFile(caminho_lote, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for nome_xml, conteudo in lote:
                    zf.writestr(c1_gerar_nome_unico(nome_xml, nomes_matriz), conteudo)

            total_xmls += len(lote)
            total_lotes += 1
            print(f"[OK] Matriz '{matriz_nome}' -> criado '{caminho_lote.name}' com {len(lote)} XML(s).")

    if total_xmls == 0:
        print("[INFO] Nenhum XML encontrado.")
    else:
        print(f"[FIM] {total_xmls} XML(s) separados em {total_lotes} lote(s).")


# ==========================================================
# Ceifador 2 — Lotes por Excel (nota+série) filtrando XML
# ==========================================================

TAG_NNF_RE = re.compile(rb"<(?:\w+:)?nNF>\s*(\d+)\s*</(?:\w+:)?nNF>")
TAG_SERIE_RE = re.compile(rb"<(?:\w+:)?serie>\s*(\d+)\s*</(?:\w+:)?serie>")


def c2_normalizar_texto(texto: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", texto)
    sem_acento = "".join(ch for ch in sem_acento if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]", "", sem_acento.lower())


def c2_valor_int(celula_valor) -> Optional[int]:
    if celula_valor is None:
        return None
    texto = str(celula_valor).strip()
    if not texto:
        return None
    digitos = re.sub(r"\D", "", texto)
    return int(digitos) if digitos else None


def c2_descobrir_colunas(cabecalho: List[str]) -> Dict[str, int]:
    mapa: Dict[str, int] = {}
    for idx, nome_coluna in enumerate(cabecalho):
        chave = c2_normalizar_texto(nome_coluna)
        if not chave:
            continue

        if "serie" in chave and "serie" not in mapa:
            mapa["serie"] = idx
        elif ("nota" in chave or "nnf" in chave or "inicial" in chave or "inicio" in chave) and (
            "inicial" in chave or "inicio" in chave or chave.endswith("ini")
        ):
            mapa.setdefault("nota_inicial", idx)
        elif ("nota" in chave or "nnf" in chave or "final" in chave or "fim" in chave) and (
            "final" in chave or "fim" in chave
        ):
            mapa.setdefault("nota_final", idx)
        elif ("nota" in chave or "nnf" in chave) and "nota" not in mapa:
            mapa["nota"] = idx
    return mapa


def c2_carregar_alvos_excel(caminho_excel: Path) -> Set[Tuple[int, int]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Dependência ausente: openpyxl") from exc

    wb = load_workbook(filename=str(caminho_excel), read_only=True, data_only=True)
    ws = wb.active

    alvos: Set[Tuple[int, int]] = set()
    encontrou_cabecalho = False
    colunas: Dict[str, int] = {}

    for linha in ws.iter_rows(values_only=True):
        valores = ["" if v is None else str(v).strip() for v in linha]
        if not encontrou_cabecalho:
            tentativa = c2_descobrir_colunas(valores)
            if "serie" in tentativa and (
                "nota" in tentativa or ("nota_inicial" in tentativa and "nota_final" in tentativa)
            ):
                colunas = tentativa
                encontrou_cabecalho = True
            continue

        serie = c2_valor_int(linha[colunas["serie"]])
        if serie is None:
            continue

        if "nota" in colunas:
            nota = c2_valor_int(linha[colunas["nota"]])
            if nota is not None:
                alvos.add((serie, nota))
            continue

        ini = c2_valor_int(linha[colunas["nota_inicial"]])
        fim = c2_valor_int(linha[colunas["nota_final"]])
        if ini is None or fim is None:
            continue
        if fim < ini:
            ini, fim = fim, ini
        for nota in range(ini, fim + 1):
            alvos.add((serie, nota))

    wb.close()

    if not encontrou_cabecalho:
        raise RuntimeError("Não encontrei cabeçalho válido no Excel (preciso de série e nota).")
    if not alvos:
        raise RuntimeError("Nenhuma combinação série+nota encontrada no Excel.")
    return alvos


def c2_extrair_serie_nota(xml_bytes: bytes) -> Optional[Tuple[int, int]]:
    m_nota = TAG_NNF_RE.search(xml_bytes)
    m_serie = TAG_SERIE_RE.search(xml_bytes)
    if not m_nota or not m_serie:
        return None
    return int(m_serie.group(1)), int(m_nota.group(1))


def ceifador2_criar_lotes_filtrados_por_excel(
    pasta_raiz: Path,
    caminho_excel: Path,
    max_xmls_por_lote: int = 500,
    prefixo_lote: str = "lote",
) -> None:
    print(f"[INFO] Lendo lista alvo no Excel: '{caminho_excel}'")
    alvos = c2_carregar_alvos_excel(caminho_excel)
    print(f"[INFO] Total de série+nota carregados: {len(alvos)}")
    print(f"[INFO] Varrendo XMLs em '{pasta_raiz}'...")

    por_matriz: Dict[str, List[Tuple[str, str, bytes, int, int]]] = {}
    encontrados: Set[Tuple[int, int]] = set()

    for matriz_nome, chave_ord, nome_xml, conteudo_xml in c1_iterar_xmls_da_pasta(pasta_raiz):
        serie_nota = c2_extrair_serie_nota(conteudo_xml)
        if serie_nota is None or serie_nota not in alvos:
            continue
        serie, nota = serie_nota
        encontrados.add((serie, nota))
        por_matriz.setdefault(matriz_nome, []).append((chave_ord, nome_xml, conteudo_xml, serie, nota))

    total_xmls = 0
    total_lotes = 0

    for matriz_nome in sorted(por_matriz.keys(), key=c1_chave_natural):
        itens = por_matriz[matriz_nome]
        itens.sort(key=lambda item: c1_chave_natural(item[0]))
        nomes_matriz: Set[str] = set()
        matriz_limpa = c1_limpar_nome_arquivo(matriz_nome)

        base = ((f"s{serie}_n{nota}_{nome_xml}", conteudo) for _, nome_xml, conteudo, serie, nota in itens)
        for idx_lote, lote in enumerate(c1_chunks(base, max_xmls_por_lote), start=1):
            caminho_lote = pasta_raiz / f"{matriz_limpa}_localizador_{prefixo_lote}{idx_lote}.zip"
            with zipfile.ZipFile(caminho_lote, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for nome_sugerido, conteudo in lote:
                    zf.writestr(c1_gerar_nome_unico(nome_sugerido, nomes_matriz), conteudo)

            total_xmls += len(lote)
            total_lotes += 1
            print(f"[OK] Matriz '{matriz_nome}' -> criado '{caminho_lote.name}' com {len(lote)} XML(s).")

    faltantes = alvos - encontrados
    if faltantes:
        print(f"[AVISO] {len(faltantes)} série+nota da planilha não foram encontradas nos XMLs.")
    else:
        print("[INFO] Todos os série+nota da planilha foram encontrados.")

    if total_xmls == 0:
        print("[INFO] Nenhum XML correspondente à planilha foi encontrado.")
    else:
        print(f"[FIM] {total_xmls} XML(s) filtrados em {total_lotes} lote(s).")


# ==========================================================
# Ceifador 3 — Validar OK e mover duplicadas sem OK
# ==========================================================


PREFIXOS_PADRAO = ("ok",)
TAMANHO_BLOCO = 1024 * 1024


@dataclass
class ArquivoInfo:
    caminho: Path
    nome: str
    extensao: str
    tamanho: int
    nome_normalizado: str
    hash_sha256: str


def c3_remover_acentos(texto: str) -> str:
    decomposed = unicodedata.normalize("NFKD", texto)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def c3_normalizar_nome_base(nome_arquivo: str, prefixos_remover: Tuple[str, ...]) -> str:
    stem = Path(nome_arquivo).stem
    texto = c3_remover_acentos(stem).lower().strip()
    texto = re.sub(r"[_\-\.\s]+", " ", texto).strip()

    alterou = True
    while alterou:
        alterou = False
        for prefixo in prefixos_remover:
            p = prefixo.strip().lower()
            if p and texto.startswith(p + " "):
                texto = texto[len(p) + 1 :].strip()
                alterou = True
            elif p and texto == p:
                texto = ""
                alterou = True

    texto = re.sub(r"[^a-z0-9 ]+", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def c3_nome_tem_prefixo(nome_arquivo: str, prefixos: Tuple[str, ...]) -> bool:
    stem = c3_remover_acentos(Path(nome_arquivo).stem).lower().strip()
    stem = re.sub(r"[_\-\.\s]+", " ", stem).strip()
    for p in prefixos:
        p_limpo = p.strip().lower()
        if not p_limpo:
            continue
        if stem == p_limpo or stem.startswith(p_limpo + " "):
            return True
    return False


def c3_calcular_sha256(caminho: Path) -> str:
    h = hashlib.sha256()
    with caminho.open("rb") as f:
        while True:
            bloco = f.read(TAMANHO_BLOCO)
            if not bloco:
                break
            h.update(bloco)
    return h.hexdigest()


def c3_coletar_arquivos(pasta_raiz: Path, prefixos_remover: Tuple[str, ...]) -> List[ArquivoInfo]:
    arquivos: List[ArquivoInfo] = []
    for caminho in pasta_raiz.rglob("*"):
        if not caminho.is_file():
            continue
        tamanho = caminho.stat().st_size
        nome = caminho.name
        extensao = caminho.suffix.lower()
        nome_normalizado = c3_normalizar_nome_base(nome, prefixos_remover)
        hash_sha256 = c3_calcular_sha256(caminho)
        arquivos.append(
            ArquivoInfo(
                caminho=caminho,
                nome=nome,
                extensao=extensao,
                tamanho=tamanho,
                nome_normalizado=nome_normalizado,
                hash_sha256=hash_sha256,
            )
        )
    return arquivos


def c3_pares_do_grupo(grupo: List[ArquivoInfo]) -> List[Tuple[ArquivoInfo, ArquivoInfo]]:
    return [(grupo[i], grupo[j]) for i in range(len(grupo)) for j in range(i + 1, len(grupo))]


def c3_validar_semelhantes(arquivos: List[ArquivoInfo], limiar_semelhanca: float) -> List[Dict[str, str]]:
    resultado: List[Dict[str, str]] = []
    por_extensao: Dict[str, List[ArquivoInfo]] = {}
    for arq in arquivos:
        por_extensao.setdefault(arq.extensao, []).append(arq)

    for extensao, grupo_ext in por_extensao.items():
        por_nome_norm: Dict[str, List[ArquivoInfo]] = {}
        for arq in grupo_ext:
            por_nome_norm.setdefault(arq.nome_normalizado, []).append(arq)

        for nome_norm, grupo in por_nome_norm.items():
            if len(grupo) < 2:
                continue
            for a, b in c3_pares_do_grupo(grupo):
                mesmo_hash = a.hash_sha256 == b.hash_sha256
                resultado.append(
                    {
                        "tipo_comparacao": "nome_normalizado_igual",
                        "extensao": extensao,
                        "nome_normalizado": nome_norm,
                        "arquivo_a": str(a.caminho),
                        "arquivo_b": str(b.caminho),
                        "tamanho_a": str(a.tamanho),
                        "tamanho_b": str(b.tamanho),
                        "hash_a": a.hash_sha256,
                        "hash_b": b.hash_sha256,
                        "mesmo_arquivo": "SIM" if mesmo_hash else "NAO",
                        "similaridade_nomes": "1.0000",
                    }
                )

        grupo_ordenado = sorted(grupo_ext, key=lambda x: x.nome.lower())
        for i in range(len(grupo_ordenado)):
            for j in range(i + 1, len(grupo_ordenado)):
                a = grupo_ordenado[i]
                b = grupo_ordenado[j]
                if a.nome_normalizado == b.nome_normalizado:
                    continue
                if a.tamanho != b.tamanho:
                    continue
                similaridade = SequenceMatcher(None, a.nome_normalizado, b.nome_normalizado).ratio()
                if similaridade < limiar_semelhanca:
                    continue
                mesmo_hash = a.hash_sha256 == b.hash_sha256
                resultado.append(
                    {
                        "tipo_comparacao": "nome_semelhante",
                        "extensao": extensao,
                        "nome_normalizado": f"{a.nome_normalizado} <-> {b.nome_normalizado}",
                        "arquivo_a": str(a.caminho),
                        "arquivo_b": str(b.caminho),
                        "tamanho_a": str(a.tamanho),
                        "tamanho_b": str(b.tamanho),
                        "hash_a": a.hash_sha256,
                        "hash_b": b.hash_sha256,
                        "mesmo_arquivo": "SIM" if mesmo_hash else "NAO",
                        "similaridade_nomes": f"{similaridade:.4f}",
                    }
                )
    return resultado


def c3_salvar_relatorio_csv(caminho_csv: Path, linhas: List[Dict[str, str]]) -> None:
    colunas = [
        "tipo_comparacao",
        "extensao",
        "nome_normalizado",
        "arquivo_a",
        "arquivo_b",
        "tamanho_a",
        "tamanho_b",
        "hash_a",
        "hash_b",
        "mesmo_arquivo",
        "similaridade_nomes",
    ]
    with caminho_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=colunas, delimiter=";")
        writer.writeheader()
        for linha in linhas:
            writer.writerow(linha)


def c3_escolher_arquivo_sem_ok(a: Path, b: Path, prefixos: Tuple[str, ...]) -> Optional[Path]:
    a_ok = c3_nome_tem_prefixo(a.name, prefixos)
    b_ok = c3_nome_tem_prefixo(b.name, prefixos)
    if a_ok and not b_ok:
        return b
    if b_ok and not a_ok:
        return a
    return None


def c3_caminho_destino_unico(destino_dir: Path, nome_arquivo: str) -> Path:
    destino = destino_dir / nome_arquivo
    if not destino.exists():
        return destino
    stem = Path(nome_arquivo).stem
    suffix = Path(nome_arquivo).suffix
    i = 1
    while True:
        cand = destino_dir / f"{stem}_{i}{suffix}"
        if not cand.exists():
            return cand
        i += 1


def c3_mover_duplicados_sem_ok(
    linhas_relatorio: List[Dict[str, str]], pasta_raiz: Path, prefixos: Tuple[str, ...]
) -> int:
    destino_dir = pasta_raiz / "duplicadas_sem_ok"
    destino_dir.mkdir(exist_ok=True)

    candidatos: Set[Path] = set()
    for linha in linhas_relatorio:
        if linha.get("mesmo_arquivo") != "SIM":
            continue
        arq_a = Path(linha["arquivo_a"])
        arq_b = Path(linha["arquivo_b"])
        sem_ok = c3_escolher_arquivo_sem_ok(arq_a, arq_b, prefixos)
        if sem_ok is not None and sem_ok.exists():
            candidatos.add(sem_ok)

    movidos = 0
    for origem in sorted(candidatos, key=lambda p: str(p).lower()):
        if not origem.exists():
            continue
        destino = c3_caminho_destino_unico(destino_dir, origem.name)
        shutil.move(str(origem), str(destino))
        movidos += 1
    return movidos


st.set_page_config(page_title="Ceifador", layout="wide")

st.title("Ceifador")

st.sidebar.header("Escolha a função")
modo = st.sidebar.radio(
    "Ceifadores",
    (
        "Ceifador 1 — Lotes por Matriz (geral)",
        "Ceifador 2 — Lotes por Excel (nota e série)",
        "Ceifador 3 — Validar OK (mover duplicadas sem OK)",
    ),
)


def bloco_saida(texto: str):
    st.subheader("Saída")
    st.code(texto or "(sem saída)", language="text")


if modo.startswith("Ceifador 1"):
    st.subheader("Ceifador 1 — Lotes por Matriz (geral)")
    pasta = st.text_input("Pasta raiz (pode ser caminho de rede)", value="")
    col1, col2 = st.columns(2)
    with col1:
        max_por_lote = st.number_input("Máximo por lote", min_value=1, value=5000, step=1)
    with col2:
        prefixo = st.text_input("Prefixo do lote (ex.: lote)", value="lote")

    if st.button("Executar Ceifador 1", type="primary", disabled=not pasta.strip()):
        try:
            buf = io.StringIO()
            with redirect_stdout(buf):
                ceifador1_criar_lotes_zip(Path(pasta.strip()), int(max_por_lote), prefixo.strip() or "lote")
            saida = buf.getvalue()
            bloco_saida(saida)
        except Exception as exc:
            st.error(str(exc))


elif modo.startswith("Ceifador 2"):
    st.subheader("Ceifador 2 — Lotes por Excel (nota e série)")
    pasta = st.text_input("Pasta raiz (pode ser caminho de rede)", value="")

    st.markdown("**Excel**: você pode informar o caminho ou fazer upload.")
    excel_caminho = st.text_input("Caminho do Excel (.xlsx)", value="")
    excel_upload = st.file_uploader("Upload do Excel (.xlsx)", type=["xlsx"])

    col1, col2 = st.columns(2)
    with col1:
        max_por_lote = st.number_input("Máximo por lote", min_value=1, value=500, step=1)
    with col2:
        prefixo = st.text_input("Prefixo do lote (ex.: lote)", value="lote")

    pode_executar = bool(pasta.strip()) and (bool(excel_caminho.strip()) or excel_upload is not None)

    if st.button("Executar Ceifador 2", type="primary", disabled=not pode_executar):
        try:
            caminho_excel_final = None
            tmp = None
            if excel_upload is not None:
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                tmp.write(excel_upload.getbuffer())
                tmp.flush()
                tmp.close()
                caminho_excel_final = tmp.name
            else:
                caminho_excel_final = excel_caminho.strip()

            buf = io.StringIO()
            with redirect_stdout(buf):
                ceifador2_criar_lotes_filtrados_por_excel(
                    pasta_raiz=Path(pasta.strip()),
                    caminho_excel=Path(caminho_excel_final),
                    max_xmls_por_lote=int(max_por_lote),
                    prefixo_lote=prefixo.strip() or "lote",
                )
            saida = buf.getvalue()
            bloco_saida(saida)
        except Exception as exc:
            st.error(str(exc))


else:
    st.subheader("Ceifador 3 — Validar OK (mover duplicadas sem OK)")
    pasta = st.text_input("Pasta raiz (pode ser caminho de rede)", value="")

    col1, col2, col3 = st.columns(3)
    with col1:
        limiar = st.number_input("Limiar de semelhança", min_value=0.0, max_value=1.0, value=0.90, step=0.01)
    with col2:
        prefixos = st.text_input("Prefixos para ignorar (separados por vírgula)", value="ok")
    with col3:
        acao = st.selectbox("Ação", options=["mover_sem_ok", "relatorio"], index=0)

    st.caption("Ação `mover_sem_ok`: mantém o arquivo com OK e move a duplicada sem OK para `duplicadas_sem_ok`.")

    if st.button("Executar Ceifador 3", type="primary", disabled=not pasta.strip()):
        try:
            prefixos_lista = [p.strip().lower() for p in prefixos.split(",") if p.strip()]
            prefixos_tuple = tuple(prefixos_lista) if prefixos_lista else v3.PREFIXOS_PADRAO

            buf = io.StringIO()
            with redirect_stdout(buf):
                pasta_raiz = Path(pasta.strip())
                arquivos = c3_coletar_arquivos(pasta_raiz, prefixos_tuple)
                linhas_relatorio = c3_validar_semelhantes(arquivos, float(limiar))
                c3_salvar_relatorio_csv(pasta_raiz / "ceifador_v3_relatorio_semelhantes.csv", linhas_relatorio)
                if acao == "mover_sem_ok":
                    c3_mover_duplicados_sem_ok(linhas_relatorio, pasta_raiz, prefixos_tuple)
            saida = buf.getvalue() or "[OK] Finalizado."
            bloco_saida(saida)
        except Exception as exc:
            st.error(str(exc))

